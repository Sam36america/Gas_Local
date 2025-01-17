import os
import glob
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
import shutil
import re
import PyPDF2

DIST = 'GÁS LOCAL'   #XML
namespace = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

class ExtratorFaturas:
    def __init__(self):
        self.regexes = {
            'cnpj': r'CPF(\s\d+\.\d+\.\d+\/\d+\-\d+)\s[E-e]',  #20/09
            'valor_total': r'\s(\d+?.?\d+.?\d+\,\d{2})\s', #20/09 (RC)
            'volume_total': r'M3?\s?\s(\d+?.?\d+.?\d+\,\d{4})',  #20/09 
            'data_emissao': r'EMISSÃO\s(\d+\/\d+\/\d+)',  #20/09
            'data_inicio': r'\:\s?(\d+\/\d+\/\d+)',  #20/09
            'data_fim': r'[A-a]\s(\d+\/\d+\/\d+)\s?', #20/09          
            'numero_fatura': r'\s(\d+)\s?\SÉRIE?', #20/09
            'valor_icms': r'(\d+?.?\d+.?\d+\,\d{2})\s?BASE' #20/09
        }

    def extrair_informacoes(self, texto):
        informacoes = {}
        for chave, regex in self.regexes.items():
            if chave == 'volume_total':
                matches = re.findall(regex, texto)
                if matches:
                    volume_total = sum([float(match.replace('.', '').replace(',', '.')) for match in matches])
                    informacoes[chave] = volume_total
            else:
                match = re.search(regex, texto)
                if match:
                    informacoes[chave] = match.group(1) if match.groups() else match.group(0)
        return informacoes

def extrair_texto(caminho_do_pdf):
    texto = ''
    with open(caminho_do_pdf, 'rb') as arquivo:
        leitor_pdf = PyPDF2.PdfReader(arquivo)
        for pagina in leitor_pdf.pages:
            texto_pagina = pagina.extract_text()
            if texto_pagina:
                texto_pagina = texto_pagina.replace('\n', ' ')
                texto += texto_pagina + ' '
    
    if not texto:
        print(f"Erro ao extrair texto do PDF: {caminho_do_pdf}")
    else:
        print(f"Texto extraído do PDF {caminho_do_pdf}: {texto[:500]}...")  # Mostra os primeiros 500 caracteres do texto extraído
    return texto.strip()  # Remove espaços extras no início e no fim

def registro_existe(df, cnpj, data_inicio, data_fim, valor_total):
    return not df[(df['CNPJ'] == cnpj) & (df['DATA INICIO'] == data_inicio) & (df['DATA FIM'] == data_fim) & (df['VALOR TOTAL'] == valor_total)].empty

def todos_campos_preenchidos(informacoes):
    campos_obrigatorios = ['cnpj', 'valor_total', 'volume_total', 'data_emissao', 'data_inicio', 'data_fim', 'numero_fatura', 'valor_icms']
    for campo in campos_obrigatorios:
        if campo not in informacoes or not informacoes[campo]:
            print(f"Campo obrigatório '{campo}' está faltando ou vazio.")
            return False
    return True

def adicionar_na_planilha(informacoes, caminho_planilha, nome_arquivo):
    if not todos_campos_preenchidos(informacoes):
        print("Não foi possível adicionar à planilha devido a campos faltantes ou vazios.")
        return False

    try:
        df = pd.read_excel(caminho_planilha)
    except FileNotFoundError:
        print(f"O arquivo '{caminho_planilha}' não foi encontrado. Criando um novo.")
        df = pd.DataFrame(columns=['CNPJ', 'VALOR TOTAL', 'VOLUME TOTAL', 'DATA EMISSAO', 'DATA INICIO', 'DATA FIM', 'NUMERO FATURA', 'VALOR ICMS', 'DISTRIBUIDORA', 'NOME DO ARQUIVO'])
    
    cnpj = informacoes['cnpj']
    data_inicio = informacoes['data_inicio']
    data_fim = informacoes['data_fim']
    valor_total = pd.to_numeric(informacoes['valor_total'].replace('.', '').replace(',', '.'))
    volume_total = informacoes['volume_total'] 
    valor_icms = pd.to_numeric(informacoes['valor_icms'].replace('.', '').replace(',', '.'))

    if registro_existe(df, cnpj, data_inicio, data_fim, valor_total):
        print(f"Registro duplicado encontrado. Não será inserido.")
        return False 
    
    
    nova_linha = pd.DataFrame([{
        'CNPJ': cnpj,
        'VALOR TOTAL': valor_total,
        'VOLUME TOTAL': volume_total,
        'DATA EMISSAO': informacoes['data_emissao'],
        'DATA INICIO': data_inicio,
        'DATA FIM': data_fim,
        'NUMERO FATURA': informacoes['numero_fatura'],
        'VALOR ICMS': valor_icms,
        'DISTRIBUIDORA': DIST,
        'NOME DO ARQUIVO': nome_arquivo
    }])
    df = pd.concat([df, nova_linha], ignore_index=True)
    df.to_excel(caminho_planilha, index=False)
    print("Dados adicionados com sucesso à planilha.")
    return True

def mover_arquivo(origem, destino):
    shutil.move(origem, destino)
    print(f"Arquivo movido para {destino}")

def verificar_linha_preenchida(caminho_planilha, informacoes):
    try:
        workbook = load_workbook(caminho_planilha)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignora o cabeçalho
            if (
                row[0] == informacoes.get('cnpj') and
                row[1] == informacoes.get('valor_total') and
                row[2] == informacoes.get('volume_total') and
                row[3] == informacoes.get('data_emissao') and
                row[4] == informacoes.get('data_inicio') and
                row[5] == informacoes.get('data_fim') and
                row[6] == informacoes.get('numero_fatura') and
                row[7] == informacoes.get('valor_icms')

            ):
                if all(cell is not None and cell != '' for cell in row):
                    return True
                else:
                    return False
        return False  # Retorna False se a linha correspondente não for encontrada
    except Exception as e:
        print(f"Erro ao verificar a planilha: {e}")
        return False

def extrair_informacoes_xml(xml_root):
    # Extrair e formatar a data de emissão
    data_emissao_iso = xml_root.find('.//nfe:ide/nfe:dhEmi', namespace).text
    data_emissao = datetime.fromisoformat(data_emissao_iso)
    data_emissao_formatada = data_emissao.strftime('%d/%m/%Y')
    
    # Clonar a data de emissão para data_inicio
    data_inicio = data_emissao_formatada
    
    # Calcular o último dia do mês para data_fim
    next_month = data_emissao.replace(day=28) + timedelta(days=4)
    data_fim = (next_month - timedelta(days=next_month.day)).strftime('%d/%m/%Y')
    
    informacoes = {
        'cnpj': xml_root.find('.//nfe:dest/nfe:CNPJ', namespace).text,
        'valor_total': xml_root.find('.//nfe:total/nfe:ICMSTot/nfe:vNF', namespace).text,
        'volume_total': xml_root.find('.//nfe:det/nfe:prod/nfe:qCom', namespace).text,
        'data_emissao': data_emissao_formatada,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'numero_fatura': xml_root.find('.//nfe:ide/nfe:nNF', namespace).text,
        'correcao_pcs': xml_root.find('.//nfe:infAdic/nfe:infCpl', namespace).text.split('Fator de Correcao Poder Calorifico: ')[1].split(' ')[0],
        'valor_icms': xml_root.find('.//nfe:total/nfe:ICMSTot/nfe:vICMS', namespace).text
    }
    return informacoes

def main(file_path, pdf_file, caminho_planilha):
    if pdf_file.endswith('.pdf') or pdf_file.endswith('.PDF'):
        texto_pypdf = extrair_texto(pdf_file)
        if not texto_pypdf:
            print(f"Erro ao extrair texto do PDF: {pdf_file}")
            return

        extrator = ExtratorFaturas()
        informacoes = extrator.extrair_informacoes(texto_pypdf)
        if not any(informacoes.values()):
            print(f"Nenhuma informação extraída do PDF: {pdf_file}")
            return
    elif pdf_file.endswith('.xml') or pdf_file.endswith('.XML'):
        tree = ET.parse(pdf_file)
        root = tree.getroot()
        informacoes = extrair_informacoes_xml(root)
    else:
        print(f"Formato de arquivo não suportado: {pdf_file}")
        return

    nome_arquivo = os.path.basename(pdf_file)
    inserido = adicionar_na_planilha(informacoes, caminho_planilha, nome_arquivo)
    print(informacoes)

    if inserido:
        destino = os.path.join(diretorio_destino, nome_arquivo)
        mover_arquivo(pdf_file, destino)

        # Mover o PDF correspondente ao XML
        base_nome_arquivo = os.path.splitext(nome_arquivo)[0]
        pdf_correspondente = os.path.join(os.path.dirname(pdf_file), base_nome_arquivo + '.pdf')
        if os.path.exists(pdf_correspondente):
            destino_pdf = os.path.join(diretorio_destino, base_nome_arquivo + '.pdf')
            mover_arquivo(pdf_correspondente, destino_pdf)
        else:
            print(f"PDF correspondente não encontrado: {pdf_correspondente}")
    else:
        print('Arquivo não foi inserido na planilha devido a dados faltantes ou duplicados. Não será movido.')

# Exemplo de uso
file_path = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\Gás Local\Faturas'
diretorio_destino = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\Gás Local\Lidos'
caminho_planilha = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\00 Faturas Lidas\GASLOCAL.xlsx'

for arquivo in os.listdir(file_path):
    if arquivo.endswith('.pdf') or arquivo.endswith('.PDF') or arquivo.endswith('.xml') or arquivo.endswith('.XML'):
        arquivo_full = os.path.join(file_path, arquivo)
        arquivo = os.path.basename(arquivo)

        main(arquivo, arquivo_full, caminho_planilha)